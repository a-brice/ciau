"""
Management command: generate the initial_data.json fixture from the XLSX file.
Usage: python manage.py generate_fixture --xlsx "path/to/CIAU Projets 2026 2.xlsx"
"""

import json
import re
from datetime import datetime, timezone
from pathlib import Path

from django.core.management.base import BaseCommand


PHASE_MAP = {
    'etudes preliminaires': 'EP',
    'avant-projet sommaire': 'APS',
    'avant-projet detaille': 'APD',
    'avant-projet d': 'APD',
    'add': 'APD',
    'instruction du permis': 'IPC',
    'projet de conception': 'PCG',
    'dossier de consultation': 'DCE',
    'dce - dqe': 'DCE',
    'mise au point': 'MDT',
    'travaux': 'TRAVAUX',
}

STAKEHOLDERS = [
    {'nom': 'Eric KOUMADO',      'poste': 'Responsable Administratif et Financier'},
    {'nom': 'Cédric GBAGBA',     'poste': 'Responsable Technique Architecture'},
    {'nom': 'Augustin AKOUVI',   'poste': 'Responsable Technique Œuvres Secondaires'},
    {'nom': 'Pierrette TCHINDI', 'poste': 'Logistique & Divers'},
]


def clean(v):
    if v is None:
        return ''
    return str(v).strip()


def normalize(text):
    """Lowercase + strip accents for matching."""
    t = clean(text).lower()
    for src, dst in [('é','e'),('è','e'),('ê','e'),('û','u'),('â','a'),('ô','o'),('\uf0a7',''),('\xa0',' ')]:
        t = t.replace(src, dst)
    return t.strip()


def parse_phase(text):
    n = normalize(text)
    for key, code in PHASE_MAP.items():
        if key in n:
            return code
    return None


def map_etat(v):
    n = normalize(v)
    if 'cours' in n:
        return 'en_cours'
    if 'dormant' in n:
        return 'dormant'
    return 'en_cours'


def map_contrat(v):
    n = normalize(v)
    if 'sign' in n:
        return 'signe'
    if 'verifier' in n or 'vérifier' in n:
        return 'a_verifier'
    if 'projet' in n:
        return 'projet'
    return ''


class Command(BaseCommand):
    help = 'Generate projects/fixtures/initial_data.json from the XLSX source file.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--xlsx',
            default='../CIAU Projets 2026 2.xlsx',
            help='Path to the source XLSX file',
        )
        parser.add_argument(
            '--out',
            default='projects/fixtures/initial_data.json',
            help='Output fixture path',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write('openpyxl is required: pip install openpyxl')
            return

        xlsx_path = Path(options['xlsx'])
        if not xlsx_path.exists():
            self.stderr.write(f'File not found: {xlsx_path}')
            return

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        projects = []
        proj_pk = 1

        # ── Liste de projets ─────────────────────────────────────────────────
        ws = wb['Liste de projets ']
        for row in ws.iter_rows(min_row=5, max_row=60, values_only=True):
            item, ref, desig, moa, contrat, etat, phase, potentiel = (list(row) + [None]*8)[:8]
            try:
                item = int(item)
            except (TypeError, ValueError):
                continue
            if not desig:
                continue
            now = datetime.now(timezone.utc).isoformat()
            projects.append({
                '_pk': proj_pk,
                'reference': clean(ref) or f'REF-{item:03d}',
                'designation': clean(desig),
                'maitre_ouvrage': clean(moa),
                'description': '',
                'responsable_etudes': '',
                'consultant_ext': '',
                'contrat_moe': map_contrat(contrat),
                'etat': map_etat(etat),
                'phase_actuelle': '',
                'potentiel': False,
                'honoraires': None,
                'created_at': now,
                'updated_at': now,
            })
            proj_pk += 1

        # ── Archives ─────────────────────────────────────────────────────────
        ws2 = wb['Archives']
        for row in ws2.iter_rows(min_row=3, max_row=40, values_only=True):
            ref = clean(row[1]) if len(row) > 1 else ''
            desig = clean(row[2]) if len(row) > 2 else ''
            moa = clean(row[3]) if len(row) > 3 else ''
            if not ref or not desig:
                continue
            if ref in ('REFERENCE', '#REF!', 'REFERENCES'):
                continue
            now = datetime.now(timezone.utc).isoformat()
            projects.append({
                '_pk': proj_pk,
                'reference': ref,
                'designation': desig,
                'maitre_ouvrage': moa,
                'description': '',
                'responsable_etudes': '',
                'consultant_ext': '',
                'contrat_moe': '',
                'etat': 'archive',
                'phase_actuelle': '',
                'potentiel': False,
                'honoraires': None,
                'created_at': now,
                'updated_at': now,
            })
            proj_pk += 1

        # ref → pk lookup
        ref_to_pk = {p['reference']: p['_pk'] for p in projects}

        # ── Fiche de Suivi → deliverables + payments ─────────────────────────
        ws3 = wb['Fiche de Suivi']
        all_rows = list(ws3.iter_rows(min_row=1, max_row=400, values_only=True))

        deliverables = []
        payments = []
        d_pk = 1
        p_pk = 1

        current_proj_pk = None
        in_livrables = False
        payment_num = 0

        for row in all_rows:
            # Project header: col[3] == 'REFERENCE :'
            if len(row) > 4 and clean(row[3]) == 'REFERENCE :' and row[4]:
                ref = clean(row[4])
                current_proj_pk = ref_to_pk.get(ref)
                in_livrables = False
                payment_num = 0
                continue

            # Identity rows
            if current_proj_pk and len(row) > 3:
                cell1 = clean(row[1])
                if "Maitre d'Ouvrage" in cell1 and row[3] and clean(row[3]) not in ('#REF!', ''):
                    for p in projects:
                        if p['_pk'] == current_proj_pk and not p['maitre_ouvrage']:
                            p['maitre_ouvrage'] = clean(row[3])
                if 'Responsable des Etudes' in cell1 and row[3]:
                    for p in projects:
                        if p['_pk'] == current_proj_pk:
                            p['responsable_etudes'] = clean(row[3])
                if 'DESCRIPTIF' in cell1 and row[1]:
                    desc = clean(row[1]).replace('DESCRIPTIF DU PROJET :', '').replace('DESCRIPTIF DU PROJET', '').strip()
                    if desc:
                        for p in projects:
                            if p['_pk'] == current_proj_pk and not p['description']:
                                p['description'] = desc

            # LIVRABLES section start
            if len(row) > 1 and row[1] and 'LIVRABLES' in str(row[1]):
                in_livrables = True
                continue

            if not in_livrables or not current_proj_pk:
                continue

            # Honoraires
            if len(row) > 4 and row[4] and 'Honoraires' in str(row[4]):
                nums = re.findall(r'\d+', str(row[4]).replace(' ', '').replace('\xa0', ''))
                for n in nums:
                    if len(n) >= 6:
                        try:
                            h = int(n)
                            for p in projects:
                                if p['_pk'] == current_proj_pk:
                                    p['honoraires'] = h
                            break
                        except ValueError:
                            pass
                continue

            # Deliverable row (col[1] has phase name)
            if len(row) > 1 and row[1]:
                cell = str(row[1])
                phase = parse_phase(cell)
                if phase:
                    etat_d = 'valide' if (len(row) > 3 and clean(row[3]) == 'x') else 'a_faire'
                    comment = clean(row[3]) if (len(row) > 3 and clean(row[3]) not in ('x', '')) else ''
                    deliverables.append({
                        'pk': d_pk,
                        'project': current_proj_pk,
                        'phase': phase,
                        'etat': etat_d,
                        'commentaire': comment if comment != 'x' else '',
                        'deadline': None,
                        'file': '',
                    })
                    d_pk += 1

            # Payment row (col[4] has 'Acompte', col[5] has amount)
            if len(row) > 5 and row[4] and 'Acompte' in str(row[4]) and isinstance(row[5], (int, float)):
                label = clean(row[4])
                m = re.search(r'(\d+)', label)
                num = int(m.group(1)) if m else (payment_num + 1)
                decl = label.split(':', 1)[1].strip() if ':' in label else label
                payments.append({
                    'pk': p_pk,
                    'project': current_proj_pk,
                    'numero': num,
                    'declenchement': decl,
                    'montant': int(row[5]),
                    'date_echeance': None,
                    'date_encaissement': None,
                })
                p_pk += 1
                payment_num = num

            # End of livrables section
            if row[1] and normalize(str(row[1])) == 'travaux':
                in_livrables = False

        # ── Build fixture ─────────────────────────────────────────────────────
        fixture = []

        for idx, s in enumerate(STAKEHOLDERS, start=1):
            fixture.append({'model': 'projects.stakeholder', 'pk': idx, 'fields': s})

        for p in projects:
            pk_val = p.pop('_pk')
            fixture.append({'model': 'projects.project', 'pk': pk_val, 'fields': p})

        for d in deliverables:
            pk_val = d.pop('pk')
            fixture.append({'model': 'projects.deliverable', 'pk': pk_val, 'fields': d})

        for pay in payments:
            pk_val = pay.pop('pk')
            fixture.append({'model': 'projects.payment', 'pk': pk_val, 'fields': pay})

        # Write
        out_path = Path(options['out'])
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(fixture, ensure_ascii=False, indent=2), encoding='utf-8')

        self.stdout.write(self.style.SUCCESS(
            f'Fixture generated: {out_path}\n'
            f'  Stakeholders : {len(STAKEHOLDERS)}\n'
            f'  Projects     : {len(projects)}\n'
            f'  Deliverables : {len(deliverables)}\n'
            f'  Payments     : {len(payments)}\n'
            f'  Total records: {len(fixture)}'
        ))
